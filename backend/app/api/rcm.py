from dataclasses import dataclass
from dataclasses import field as dc_field
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import CurrentUser
from app.models.rcm import (
    Control,
    ControlAssertion,
    Process,
    Risk,
    RiskCategory,
    SubProcess,
)
from app.models.rcm_baseline import (
    ACTION_ADD,
    ACTION_ADOPT,
    ACTION_EXCLUDE,
    ACTION_OVERRIDE,
    BaselineControl,
    BaselineProcess,
    BaselineRisk,
    BaselineSubProcess,
    ControlInstance,
    ProcessInstance,
    RiskInstance,
    SubProcessInstance,
)
from app.schemas.rcm import (
    BulkDeleteRequest,
    BulkUpdateRequest,
    ControlAssertionCreate,
    ControlAssertionRead,
    ControlCreate,
    ControlRead,
    ControlSearchOut,
    ControlSearchResponse,
    ControlUpdate,
    ProcessCreate,
    ProcessRead,
    ProcessUpdate,
    RiskCategoryCreate,
    RiskCategoryRead,
    RiskCategoryUpdate,
    RiskCreate,
    RiskRead,
    RiskUpdate,
    SubProcessCreate,
    SubProcessRead,
    SubProcessUpdate,
)
from app.services.control_resolver import (
    resolve_controls,
    resolve_hierarchy,
    resolve_processes,
)
from app.services.excel_parser import find_rcm_sheet

router = APIRouter(prefix="/api/rcm", tags=["rcm"])


# ── 모듈 정보 ─────────────────────────────────────────────

@router.get("/info")
def get_module_info(user: CurrentUser) -> dict:
    return {
        "module": "rcm",
        "name_kr": "RCM 관리",
        "phase_0_status": "Phase 1 풀 확장 완료",
        "phase_1_features": [
            "SubProcess·Risk·Control CRUD",
            "Excel 업로드 (사이냅소프트 양식)",
            "풀 검색·필터",
            "위험 매트릭스",
            "벌크 삭제·수정",
        ],
        "phase_1_excluded": ["버전 관리 스냅샷·Diff", "변경 승인 워크플로"],
        "available_in_phase_1": True,
    }


# ── Processes ─────────────────────────────────────────────

# ── 상위 3계층 overlay CRUD 헬퍼 (2-A-4-3, ADR-0029) ───────
# 통제(_apply_control_update/_apply_control_delete)와 같은 규칙을 계층 모델로 매개화한 것.
# 계층마다 복붙하지 않되 클래스·서비스는 두지 않는다(ADR-0020 — _resolve_layer 선례).

# override 대상 = 각 계층 Update 스키마 필드 (code·상위참조는 제외 — 정체성/구조라 override 대상 아님)
_PROCESS_OVERRIDE_FIELDS = list(ProcessUpdate.model_fields.keys())
_SUB_PROCESS_OVERRIDE_FIELDS = list(SubProcessUpdate.model_fields.keys())
_RISK_OVERRIDE_FIELDS = list(RiskUpdate.model_fields.keys())


def _assert_code_available(db: Session, baseline_model, instance_model, code: str, label: str) -> None:
    """신규 add 의 code 중복 검증 (ADR-0029 §3).

    DB 제약은 `(tenant_id, code)` 뿐이라 **instance 끼리의 충돌만** 막는다.
    baseline 테이블의 code 와 겹치는 경우는 어떤 제약도 막지 못하므로 여기서 함께 본다
    (겹치면 resolver 결과에 같은 code 가 둘 나온다). 제약 위반을 IntegrityError 로 터뜨리는
    대신 의미 있는 409 를 돌려준다.
    """
    dup_baseline = db.query(baseline_model).filter(
        baseline_model.code == code,
        baseline_model.is_deleted == False,  # noqa: E712
    ).first()
    if dup_baseline is not None:
        raise HTTPException(status_code=409, detail=f"{label} 코드 '{code}' 는 표준(baseline)에 이미 있습니다")
    dup_inst = db.query(instance_model).filter(
        instance_model.code == code,
        instance_model.is_deleted == False,  # noqa: E712
    ).first()
    if dup_inst is not None:
        raise HTTPException(status_code=409, detail=f"{label} 코드 '{code}' 는 이미 사용 중입니다")


def _apply_layer_update(db: Session, baseline_model, instance_model, fk_attr: str,
                        override_fields: list[str], item_id: UUID, changes: dict) -> bool:
    """계층 수정 — baseline 유래는 override instance 필드 diff, add 는 instance 직접 수정.

    요청 값이 baseline 과 같으면 NULL(=baseline 따름), 다르면 값 저장. 전부 NULL 이면 adopt 로
    되돌린다(instance 는 검토 흔적으로 남긴다). **커밋하지 않는다** — 경계는 호출자.
    """
    baseline = db.query(baseline_model).filter(baseline_model.id == item_id).first()
    if baseline is not None:
        inst = db.query(instance_model).filter(getattr(instance_model, fk_attr) == item_id).first()
        if inst is None:
            inst = instance_model(**{fk_attr: item_id}, action=ACTION_OVERRIDE)
            db.add(inst)
        elif inst.action in (ACTION_ADOPT, ACTION_EXCLUDE):
            inst.action = ACTION_OVERRIDE
        for f in override_fields:
            if f in changes:  # 전송된 필드만 (None 도 전송이면 diff 대상)
                req_val = changes[f]
                setattr(inst, f, None if req_val == getattr(baseline, f) else req_val)
        if all(getattr(inst, f) is None for f in override_fields):
            inst.action = ACTION_ADOPT
        return True

    inst = db.query(instance_model).filter(
        instance_model.id == item_id,
        instance_model.action == ACTION_ADD,
        instance_model.is_deleted == False,  # noqa: E712
    ).first()
    if inst is None:
        return False
    for f, v in changes.items():
        setattr(inst, f, v)
    return True


def _apply_layer_delete(db: Session, baseline_model, instance_model, fk_attr: str,
                        override_fields: list[str], item_id: UUID, parent_attrs: tuple[str, ...]) -> bool:
    """계층 삭제 — baseline 유래는 exclude instance(원본 baseline 불변), add 는 soft delete.

    하위 계층에는 아무것도 쓰지 않는다 — "상위로 인한 제외"는 저장하지 않고 조회 시점에
    계산한다(ADR-0029 §2.2). **커밋하지 않는다** — 경계는 호출자.
    """
    baseline = db.query(baseline_model).filter(baseline_model.id == item_id).first()
    if baseline is not None:
        inst = db.query(instance_model).filter(getattr(instance_model, fk_attr) == item_id).first()
        if inst is None:
            inst = instance_model(**{fk_attr: item_id}, action=ACTION_EXCLUDE)
            db.add(inst)
        else:
            inst.action = ACTION_EXCLUDE
            for f in override_fields:  # override 흔적 정리
                setattr(inst, f, None)
            inst.code = None
            for a in parent_attrs:
                setattr(inst, a, None)
        return True

    inst = db.query(instance_model).filter(
        instance_model.id == item_id,
        instance_model.action == ACTION_ADD,
        instance_model.is_deleted == False,  # noqa: E712
    ).first()
    if inst is None:
        return False
    inst.is_deleted = True
    return True


def _resolve_process_parent(db: Session, process_id: UUID | None) -> tuple[UUID | None, UUID | None]:
    """요청 process_id → (process_baseline_id, process_instance_id). 2-B-2 이중 FK 규칙."""
    if process_id is None:
        return None, None
    if db.query(BaselineProcess).filter(BaselineProcess.id == process_id).first() is not None:
        return process_id, None
    if db.query(ProcessInstance).filter(ProcessInstance.id == process_id).first() is not None:
        return None, process_id
    return None, None


def _resolve_sub_process_parent(db: Session, sub_process_id: UUID | None) -> tuple[UUID | None, UUID | None]:
    """요청 sub_process_id → (sub_process_baseline_id, sub_process_instance_id)."""
    if sub_process_id is None:
        return None, None
    if db.query(BaselineSubProcess).filter(BaselineSubProcess.id == sub_process_id).first() is not None:
        return sub_process_id, None
    if db.query(SubProcessInstance).filter(SubProcessInstance.id == sub_process_id).first() is not None:
        return None, sub_process_id
    return None, None


# ── 상위 3계층 resolver 조회 헬퍼 (2-A-4-3, ADR-0029) ──────
# cascade 로 빠진 항목은 목록에 없으므로 상세도 404 — 조회 경로 일관.

def _resolved_process_or_404(db: Session, process_id: UUID) -> dict:
    for r in resolve_processes(db):
        if r["id"] == process_id:
            return r
    raise HTTPException(status_code=404, detail="Process not found")


def _resolved_sub_process_or_404(db: Session, sp_id: UUID) -> dict:
    for r in resolve_hierarchy(db)[1]:
        if r["id"] == sp_id:
            return r
    raise HTTPException(status_code=404, detail="SubProcess not found")


def _resolved_risk_or_404(db: Session, risk_id: UUID) -> dict:
    for r in resolve_hierarchy(db)[2]:
        if r["id"] == risk_id:
            return r
    raise HTTPException(status_code=404, detail="Risk not found")


@router.get("/processes")
def list_processes(skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """목록 — resolve_processes 경유 (2-A-4-3, ADR-0029). 최상위 계층이라 cascade 대상 없음."""
    rows = resolve_processes(db)
    rows.sort(key=lambda r: r["code"] or "")
    total = len(rows)
    page = rows[skip: skip + limit]
    return {"items": [ProcessRead(**r) for r in page], "total": total, "skip": skip, "limit": limit}


@router.post("/processes", status_code=status.HTTP_201_CREATED, response_model=ProcessRead)
def create_process(body: ProcessCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """add instance 생성 (2-A-4-3, ADR-0029). tenant_id 는 before_flush 자동 stamp."""
    _assert_code_available(db, BaselineProcess, ProcessInstance, body.code, "프로세스")
    inst = ProcessInstance(action=ACTION_ADD, baseline_process_id=None, **body.model_dump())
    db.add(inst)
    db.commit()
    return _resolved_process_or_404(db, inst.id)


@router.get("/processes/{process_id}", response_model=ProcessRead)
def get_process(process_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """상세 — resolver 결과에서 정체성 id 매칭 (2-A-4-3, ADR-0029)."""
    return _resolved_process_or_404(db, process_id)


@router.patch("/processes/{process_id}", response_model=ProcessRead)
def update_process(process_id: UUID, body: ProcessUpdate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """수정 — baseline 유래면 override instance, add 면 instance 직접 (2-A-4-3, ADR-0029).

    `exclude_unset` — False/""/None 도 유효한 값이라 미전송 여부로만 판별한다(2-A-4-2 선례).
    """
    if not _apply_layer_update(db, BaselineProcess, ProcessInstance, "baseline_process_id",
                               _PROCESS_OVERRIDE_FIELDS, process_id, body.model_dump(exclude_unset=True)):
        raise HTTPException(status_code=404, detail="Process not found")
    db.commit()
    return _resolved_process_or_404(db, process_id)


@router.delete("/processes/{process_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_process(process_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    """삭제 — baseline 유래는 exclude instance, add 는 soft delete (2-A-4-3, ADR-0029).

    하위(sub_process/risk/control)는 건드리지 않는다 — cascade 는 조회 시점 계산(§2.2).
    """
    if not _apply_layer_delete(db, BaselineProcess, ProcessInstance, "baseline_process_id",
                               _PROCESS_OVERRIDE_FIELDS, process_id, ()):
        raise HTTPException(status_code=404, detail="Process not found")
    db.commit()


# ── SubProcesses ──────────────────────────────────────────

@router.get("/sub-processes")
def list_sub_processes(process_id: UUID | None = None, skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """목록 — resolve_hierarchy 경유 (2-A-4-3, ADR-0029). 상위 process 제외분은 cascade 로 빠진다."""
    rows = resolve_hierarchy(db)[1]
    if process_id:
        rows = [r for r in rows if r["process_id"] == process_id]
    rows.sort(key=lambda r: r["code"] or "")
    total = len(rows)
    page = rows[skip: skip + limit]
    return {"items": [SubProcessRead(**r) for r in page], "total": total, "skip": skip, "limit": limit}


@router.post("/sub-processes", status_code=status.HTTP_201_CREATED, response_model=SubProcessRead)
def create_sub_process(body: SubProcessCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """add instance 생성 (2-A-4-3, ADR-0029). 상위는 이중 FK 규칙으로 baseline/instance 매핑."""
    _assert_code_available(db, BaselineSubProcess, SubProcessInstance, body.code, "하위프로세스")
    data = body.model_dump()
    pb, pi = _resolve_process_parent(db, data.pop("process_id"))
    inst = SubProcessInstance(action=ACTION_ADD, baseline_sub_process_id=None,
                              process_baseline_id=pb, process_instance_id=pi, **data)
    db.add(inst)
    db.commit()
    return _resolved_sub_process_or_404(db, inst.id)


@router.get("/sub-processes/{sp_id}", response_model=SubProcessRead)
def get_sub_process(sp_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """상세 — resolver 결과에서 정체성 id 매칭 (2-A-4-3, ADR-0029)."""
    return _resolved_sub_process_or_404(db, sp_id)


@router.patch("/sub-processes/{sp_id}", response_model=SubProcessRead)
def update_sub_process(sp_id: UUID, body: SubProcessUpdate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """수정 — baseline 유래면 override instance, add 면 instance 직접 (2-A-4-3, ADR-0029)."""
    if not _apply_layer_update(db, BaselineSubProcess, SubProcessInstance, "baseline_sub_process_id",
                               _SUB_PROCESS_OVERRIDE_FIELDS, sp_id, body.model_dump(exclude_unset=True)):
        raise HTTPException(status_code=404, detail="SubProcess not found")
    db.commit()
    return _resolved_sub_process_or_404(db, sp_id)


@router.delete("/sub-processes/{sp_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_sub_process(sp_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    """삭제 — baseline 유래는 exclude instance, add 는 soft delete (2-A-4-3, ADR-0029)."""
    if not _apply_layer_delete(db, BaselineSubProcess, SubProcessInstance, "baseline_sub_process_id",
                               _SUB_PROCESS_OVERRIDE_FIELDS, sp_id,
                               ("process_baseline_id", "process_instance_id")):
        raise HTTPException(status_code=404, detail="SubProcess not found")
    db.commit()


# ── Risks ─────────────────────────────────────────────────

@router.get("/risks")
def list_risks(sub_process_id: UUID | None = None, skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """목록 — resolve_hierarchy 경유 (2-A-4-3, ADR-0029). 상위 계층 제외분은 cascade 로 빠진다."""
    rows = resolve_hierarchy(db)[2]
    if sub_process_id:
        rows = [r for r in rows if r["sub_process_id"] == sub_process_id]
    rows.sort(key=lambda r: r["code"] or "")
    total = len(rows)
    page = rows[skip: skip + limit]
    return {"items": [RiskRead(**r) for r in page], "total": total, "skip": skip, "limit": limit}


@router.post("/risks", status_code=status.HTTP_201_CREATED, response_model=RiskRead)
def create_risk(body: RiskCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """add instance 생성 (2-A-4-3, ADR-0029). 상위는 이중 FK 규칙으로 baseline/instance 매핑."""
    _assert_code_available(db, BaselineRisk, RiskInstance, body.code, "위험")
    data = body.model_dump()
    sb, si = _resolve_sub_process_parent(db, data.pop("sub_process_id"))
    inst = RiskInstance(action=ACTION_ADD, baseline_risk_id=None,
                        sub_process_baseline_id=sb, sub_process_instance_id=si, **data)
    db.add(inst)
    db.commit()
    return _resolved_risk_or_404(db, inst.id)


@router.get("/risks/{risk_id}", response_model=RiskRead)
def get_risk(risk_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """상세 — resolver 결과에서 정체성 id 매칭 (2-A-4-3, ADR-0029)."""
    return _resolved_risk_or_404(db, risk_id)


@router.patch("/risks/{risk_id}", response_model=RiskRead)
def update_risk(risk_id: UUID, body: RiskUpdate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """수정 — baseline 유래면 override instance, add 면 instance 직접 (2-A-4-3, ADR-0029)."""
    if not _apply_layer_update(db, BaselineRisk, RiskInstance, "baseline_risk_id",
                               _RISK_OVERRIDE_FIELDS, risk_id, body.model_dump(exclude_unset=True)):
        raise HTTPException(status_code=404, detail="Risk not found")
    db.commit()
    return _resolved_risk_or_404(db, risk_id)


@router.delete("/risks/{risk_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_risk(risk_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    """삭제 — baseline 유래는 exclude instance, add 는 soft delete (2-A-4-3, ADR-0029)."""
    if not _apply_layer_delete(db, BaselineRisk, RiskInstance, "baseline_risk_id",
                               _RISK_OVERRIDE_FIELDS, risk_id,
                               ("sub_process_baseline_id", "sub_process_instance_id")):
        raise HTTPException(status_code=404, detail="Risk not found")
    db.commit()


# ── Risk Categories ───────────────────────────────────────

@router.get("/risk-categories")
def list_risk_categories(skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    q = db.query(RiskCategory).filter(RiskCategory.is_deleted == False)  # noqa: E712
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    return {"items": [RiskCategoryRead.model_validate(i) for i in items], "total": total, "skip": skip, "limit": limit}


@router.post("/risk-categories", status_code=status.HTTP_201_CREATED, response_model=RiskCategoryRead)
def create_risk_category(body: RiskCategoryCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> RiskCategory:
    obj = RiskCategory(**body.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.get("/risk-categories/{rc_id}", response_model=RiskCategoryRead)
def get_risk_category(rc_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> RiskCategory:
    obj = db.query(RiskCategory).filter(RiskCategory.id == rc_id, RiskCategory.is_deleted == False).first()  # noqa: E712
    if not obj:
        raise HTTPException(status_code=404, detail="RiskCategory not found")
    return obj


@router.patch("/risk-categories/{rc_id}", response_model=RiskCategoryRead)
def update_risk_category(rc_id: UUID, body: RiskCategoryUpdate, user: CurrentUser = None, db: Session = Depends(get_db)) -> RiskCategory:
    obj = db.query(RiskCategory).filter(RiskCategory.id == rc_id, RiskCategory.is_deleted == False).first()  # noqa: E712
    if not obj:
        raise HTTPException(status_code=404, detail="RiskCategory not found")
    for f, v in body.model_dump(exclude_none=True).items():
        setattr(obj, f, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/risk-categories/{rc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_risk_category(rc_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    obj = db.query(RiskCategory).filter(RiskCategory.id == rc_id, RiskCategory.is_deleted == False).first()  # noqa: E712
    if not obj:
        raise HTTPException(status_code=404, detail="RiskCategory not found")
    obj.is_deleted = True
    db.commit()


# ── Controls — 정적 경로 먼저 (파라미터 경로보다 앞에 위치해야 함) ──

@router.get("/controls/search", response_model=ControlSearchResponse)
def search_controls(
    q: str | None = None,
    process_code: str | None = None,
    sub_process_code: str | None = None,
    risk_level: str | None = None,
    frequency: str | None = None,
    is_key_control: bool | None = None,
    auto_manual: str | None = None,
    preventive_detective: str | None = None,
    assertion: str | None = None,
    owner: str | None = None,
    skip: int = 0,
    limit: int = 100,
    sort_by: str = "code",
    sort_order: str = "asc",
    user: CurrentUser = None,
    db: Session = Depends(get_db),
) -> dict:
    """풀 검색·필터 — resolve_controls(baseline−exclude+override+add) 위의 메모리 처리 (ADR-0027, 2-A-3).

    기존 SQL 조인/WHERE/ORDER BY 를 resolve_controls 결과 dict 위 필터·정렬로 재구현.
    관계 필드(process_code/sub_process_code/risk_level/assertions)·source envelope 는
    resolver 가 이미 채워 반환하므로 여기서 조인하지 않는다.
    """
    rows = resolve_controls(db)  # 활성 tenant 자동 격리, 관계 필드·envelope 포함

    def _has(text: str | None, needle: str) -> bool:
        return text is not None and needle.lower() in text.lower()

    filtered = []
    for r in rows:
        if q and not (_has(r["code"], q) or _has(r["name"], q)
                      or _has(r["description"], q) or _has(r["owner_name"], q)):
            continue
        if owner and not _has(r["owner_name"], owner):
            continue
        if frequency and r["frequency"] != frequency:
            continue
        if is_key_control is not None and r["is_key_control"] != is_key_control:
            continue
        if auto_manual and r["auto_manual"] != auto_manual:
            continue
        if preventive_detective and r["preventive_detective"] != preventive_detective:
            continue
        if risk_level and r["risk_level"] != risk_level:
            continue
        if sub_process_code and r["sub_process_code"] != sub_process_code:
            continue
        if process_code and r["process_code"] != process_code:
            continue
        if assertion and assertion not in r["assertions"]:
            continue
        filtered.append(r)

    valid_sort = {"code", "name", "frequency", "created_at", "owner_name"}
    key = sort_by if sort_by in valid_sort else "code"
    # None 안전 정렬 — (있음/없음, 값) 튜플로 None 을 뒤로. datetime/str 혼용 없음(키별 동일 타입).
    filtered.sort(
        key=lambda r: (r.get(key) is None, r.get(key)),
        reverse=(sort_order == "desc"),
    )

    total = len(filtered)
    page = filtered[skip: skip + limit]
    items_out = [ControlSearchOut(**r) for r in page]

    return {
        "items": items_out,
        "total": total,
        "skip": skip,
        "limit": limit,
        "sort": f"{sort_by}:{sort_order}",
    }


@router.post("/controls/bulk-delete")
def bulk_delete_controls(body: BulkDeleteRequest, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """다건 삭제 — 단건 DELETE 와 동일 분기를 id 마다 적용 (ADR-0027, 2-A-4-2).

    미해당 id 는 건너뛰고 `skipped_ids` 로 드러낸다 — 하나 때문에 전체를 404 로 실패시키지 않는다.
    단일 트랜잭션(루프 후 1회 커밋) — 중간 실패 시 전체 롤백.
    """
    deleted, skipped = 0, []
    for cid in body.control_ids:
        if _apply_control_delete(db, cid):
            deleted += 1
        else:
            skipped.append(cid)
    db.commit()
    return {"deleted_count": deleted, "skipped_ids": skipped}


@router.post("/controls/bulk-update")
def bulk_update_controls(body: BulkUpdateRequest, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """다건 수정 — 단건 PATCH 와 동일 분기를 id 마다 적용 (ADR-0027, 2-A-4-2).

    `exclude_unset` — 단건 PATCH 와 동일 기준으로 통일(2-A-4-2). 이전 `exclude_none` 은
    False/0/"" 를 미전송으로 오인해 저장할 수 없었다.
    단일 트랜잭션(루프 후 1회 커밋).
    """
    changes = body.updates.model_dump(exclude_unset=True)
    updated, skipped = 0, []
    for cid in body.control_ids:
        if _apply_control_update(db, cid, changes):
            updated += 1
        else:
            skipped.append(cid)
    db.commit()
    return {"updated_count": updated, "skipped_ids": skipped}


@router.get("/controls")
def list_controls(skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """목록 — resolve_controls 경유 (ADR-0027, 2-A-4-2).

    search·상세와 같은 정체성 id 체계 + source envelope. 정렬은 search 기본과 동일(code 오름차순).
    """
    rows = resolve_controls(db)
    rows.sort(key=lambda r: (r["code"] is None, r["code"]))
    total = len(rows)
    page = rows[skip: skip + limit]
    return {"items": [ControlRead(**r) for r in page], "total": total, "skip": skip, "limit": limit}


# PATCH override diff 대상 = ControlUpdate 필드(코드·risk_id 제외한 미러링 필드).
_OVERRIDE_FIELDS = list(ControlUpdate.model_fields.keys())


def _resolve_risk_parent(db: Session, risk_id: UUID | None) -> tuple[UUID | None, UUID | None]:
    """요청 risk_id → (risk_baseline_id, risk_instance_id). 2-B-2 이중 FK 규칙.

    baseline_risks 에 있으면 baseline 쪽, risk_instances 에 있으면 instance 쪽.
    어느 쪽도 아니면 (None, None) — 상위 없이 생성(cascade 에서 risk_id NULL 은 유지).
    둘 다 non-NULL 은 불가(check) 이므로 배타적으로 하나만 반환.
    """
    if risk_id is None:
        return None, None
    if db.query(BaselineRisk).filter(BaselineRisk.id == risk_id).first() is not None:
        return risk_id, None
    if db.query(RiskInstance).filter(RiskInstance.id == risk_id).first() is not None:
        return None, risk_id
    return None, None


def _resolved_or_404(db: Session, control_id: UUID) -> dict:
    for r in resolve_controls(db):
        if r["id"] == control_id:
            return r
    raise HTTPException(status_code=404, detail="Control not found")


def _apply_control_update(db: Session, control_id: UUID, changes: dict) -> bool:
    """수정 분기 — 단건 PATCH 와 다건 bulk-update 공통 (2-A-4-1, 2-A-4-2 추출).

    - baseline 유래 → override instance 생성/갱신 (바뀐 필드만 저장, 필드 diff).
      요청 값이 baseline 과 같으면 NULL(=baseline 따름), 다르면 값 저장.
      모든 override 필드가 NULL 이면 action 을 adopt 로 전환(instance 는 남김 — 검토 흔적).
    - 회사 add → instance 직접 수정 (baseline 없어 diff 불필요).

    False/0/"" 는 유효 값이므로 호출자가 `exclude_unset` 으로 미전송을 판별해 changes 를
    만든다(falsy 판정 금지). **커밋하지 않는다** — 트랜잭션 경계는 호출자(다건은 1회 커밋).
    대상이 없으면 False.
    """
    baseline = db.query(BaselineControl).filter(BaselineControl.id == control_id).first()
    if baseline is not None:
        inst = db.query(ControlInstance).filter(
            ControlInstance.baseline_control_id == control_id
        ).first()  # (tenant_id, baseline_control_id) unique → 최대 1건
        if inst is None:
            inst = ControlInstance(baseline_control_id=control_id, action=ACTION_OVERRIDE)
            db.add(inst)
        elif inst.action in (ACTION_ADOPT, ACTION_EXCLUDE):
            inst.action = ACTION_OVERRIDE
        for f in _OVERRIDE_FIELDS:
            if f in changes:  # 전송된 필드만 (None 도 전송이면 diff 대상)
                base_val = getattr(baseline, f)
                req_val = changes[f]
                setattr(inst, f, None if req_val == base_val else req_val)
        if all(getattr(inst, f) is None for f in _OVERRIDE_FIELDS):
            inst.action = ACTION_ADOPT  # 전부 baseline 과 동일 → 되돌림 (instance 는 남김)
        return True

    inst = db.query(ControlInstance).filter(
        ControlInstance.id == control_id,
        ControlInstance.action == ACTION_ADD,
        ControlInstance.is_deleted == False,  # noqa: E712
    ).first()
    if inst is None:
        return False
    for f, v in changes.items():
        setattr(inst, f, v)
    return True


def _apply_control_delete(db: Session, control_id: UUID) -> bool:
    """삭제 분기 — 단건 DELETE 와 다건 bulk-delete 공통 (2-A-4-1, 2-B-3.5 규약, 2-A-4-2 추출).

    - baseline 유래 → exclude instance 생성/전환 (원본 baseline_controls 절대 불변, hide).
    - 회사 add → instance soft delete.

    **커밋하지 않는다** — 트랜잭션 경계는 호출자. 대상이 없으면 False.
    """
    baseline = db.query(BaselineControl).filter(BaselineControl.id == control_id).first()
    if baseline is not None:
        inst = db.query(ControlInstance).filter(
            ControlInstance.baseline_control_id == control_id
        ).first()
        if inst is None:
            inst = ControlInstance(baseline_control_id=control_id, action=ACTION_EXCLUDE)
            db.add(inst)
        else:
            inst.action = ACTION_EXCLUDE
            for f in _OVERRIDE_FIELDS:  # override 필드 정리
                setattr(inst, f, None)
            inst.code = None
            inst.risk_baseline_id = None
            inst.risk_instance_id = None
        return True

    inst = db.query(ControlInstance).filter(
        ControlInstance.id == control_id,
        ControlInstance.action == ACTION_ADD,
        ControlInstance.is_deleted == False,  # noqa: E712
    ).first()
    if inst is None:
        return False
    inst.is_deleted = True
    return True


@router.post("/controls", status_code=status.HTTP_201_CREATED, response_model=ControlRead)
def create_control(body: ControlCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """add instance 생성 (ADR-0027, 2-A-4-1). 회사 고유 통제 = ControlInstance(action=ACTION_ADD).

    tenant_id 는 before_flush 자동 stamp(ADR-0026, 수동 지정 금지).
    상위 risk 참조는 이중 FK 규칙으로 baseline/instance 중 하나에 매핑.
    """
    data = body.model_dump()
    risk_id = data.pop("risk_id")
    rb, ri = _resolve_risk_parent(db, risk_id)
    inst = ControlInstance(
        action=ACTION_ADD, baseline_control_id=None,
        risk_baseline_id=rb, risk_instance_id=ri,
        **data,
    )
    db.add(inst)
    db.commit()
    return _resolved_or_404(db, inst.id)


@router.get("/controls/{control_id}", response_model=ControlRead)
def get_control(control_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """상세 — resolve_controls 결과에서 정체성 id 매칭 (ADR-0027, 2-A-3).

    목록(search)과 id 체계 일치 — 프론트가 목록에서 받은 id 를 그대로 사용.
    단건을 위해 전체 resolve 하는 것은 비효율이나 95~수백 규모에서 무방(구조 정합성 우선).
    """
    for r in resolve_controls(db):
        if r["id"] == control_id:
            return r  # ControlRead 로 검증 — 관계 필드 등 초과 키는 무시, envelope 포함
    raise HTTPException(status_code=404, detail="Control not found")


@router.patch("/controls/{control_id}", response_model=ControlRead)
def update_control(control_id: UUID, body: ControlUpdate, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """단건 수정 — 분기는 `_apply_control_update` 공통 (ADR-0027, 2-A-4-1).

    `exclude_unset` 으로 미전송을 판별한다 — False/0/"" 는 유효 값(falsy 판정 금지).
    """
    if not _apply_control_update(db, control_id, body.model_dump(exclude_unset=True)):
        raise HTTPException(status_code=404, detail="Control not found")
    db.commit()
    return _resolved_or_404(db, control_id)


@router.delete("/controls/{control_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_control(control_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    """단건 삭제 — 분기는 `_apply_control_delete` 공통 (ADR-0027, 2-A-4-1, 2-B-3.5 삭제 규약)."""
    if not _apply_control_delete(db, control_id):
        raise HTTPException(status_code=404, detail="Control not found")
    db.commit()


# ── Control Assertions ────────────────────────────────────

@router.get("/control-assertions")
def list_control_assertions(skip: int = 0, limit: int = 100, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    q = db.query(ControlAssertion).filter(ControlAssertion.is_deleted == False)  # noqa: E712
    total = q.count()
    items = q.offset(skip).limit(limit).all()
    return {"items": [ControlAssertionRead.model_validate(i) for i in items], "total": total, "skip": skip, "limit": limit}


@router.post("/control-assertions", status_code=status.HTTP_201_CREATED, response_model=ControlAssertionRead)
def create_control_assertion(body: ControlAssertionCreate, user: CurrentUser = None, db: Session = Depends(get_db)) -> ControlAssertion:
    obj = ControlAssertion(**body.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/control-assertions/{ca_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_control_assertion(ca_id: UUID, user: CurrentUser = None, db: Session = Depends(get_db)) -> None:
    obj = db.query(ControlAssertion).filter(ControlAssertion.id == ca_id, ControlAssertion.is_deleted == False).first()  # noqa: E712
    if not obj:
        raise HTTPException(status_code=404, detail="ControlAssertion not found")
    obj.is_deleted = True
    db.commit()


# ── Excel 업로드 ──────────────────────────────────────────

@dataclass
class _ParsedRCM:
    processes: dict
    sub_processes: dict
    risks: dict
    controls: list
    errors: list
    warnings: list = dc_field(default_factory=list)


def _parse_rcm_sheet(ws, header_row: int, mapping: dict[str, int]) -> _ParsedRCM:
    """헤더 위치가 결정된 시트에서 RCM 데이터 파싱.

    필수 3개 컬럼은 mapping 사용. 나머지는 사이냅소프트 양식 고정 위치 유지.
    """
    valid_levels = {"LR", "MR", "HR", "SR"}
    valid_freq = {"O", "D", "W", "M", "Q", "A"}
    valid_pd = {"P", "D"}
    valid_am = {"A", "M", "IT"}
    valid_ipe = {"Y", "N", "N/A"}

    p_col = mapping["process_code"]
    c_col = mapping["control_code"]
    cn_col = mapping["control_name"]

    processes: dict = {}
    sub_processes: dict = {}
    risks: dict = {}
    controls: list = []
    errors: list = []
    warnings: list = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not row or row[p_col] is None:
            break
        try:
            p_code = str(row[p_col]).strip()
            p_name = str(row[2] or "").strip()
            sp_code = str(row[3] or "").strip()
            sp_name = str(row[4] or "").strip()
            r_code = str(row[5] or "").strip()
            c_code = str(row[c_col] or "").strip()

            if not c_code:
                errors.append(f"Row {row_idx}: 통제활동번호 누락")
                continue
            if not r_code:
                errors.append(f"Row {row_idx}: 위험번호 누락")
                continue

            processes[p_code] = p_name
            sub_processes[sp_code] = {"name": sp_name, "process_code": p_code}

            level = str(row[14] or "LR").strip().upper()
            if level not in valid_levels:
                warnings.append(f"Row {row_idx}: 위험평가 '{level}' 무효 → LR 사용")
                level = "LR"
            risks[r_code] = {
                "description": str(row[8] or r_code).strip(),
                "assessment_level": level,
                "sub_process_code": sp_code,
            }

            c_name = str(row[cn_col] or "").strip()
            if not c_name:
                errors.append(f"Row {row_idx}: 통제활동이름 누락")
                continue

            pd_val = str(row[25] or "P").strip().upper()
            if pd_val not in valid_pd:
                warnings.append(f"Row {row_idx}: P/D 값 '{pd_val}' 무효 → P 사용")
                pd_val = "P"

            am_val = str(row[26] or "M").strip().upper()
            if am_val not in valid_am:
                warnings.append(f"Row {row_idx}: Auto/Manual '{am_val}' 무효 → M 사용")
                am_val = "M"

            freq_val = str(row[35] or "A").strip().upper()
            if freq_val not in valid_freq:
                warnings.append(f"Row {row_idx}: 통제주기 '{freq_val}' 무효 → A 사용")
                freq_val = "A"

            ipe_val = str(row[36] or "N/A").strip()
            if ipe_val not in valid_ipe:
                warnings.append(f"Row {row_idx}: IPE '{ipe_val}' 무효 → N/A 사용")
                ipe_val = "N/A"

            assertion_map = [
                ("E", 27), ("C", 28), ("R", 29), ("V", 30),
                ("P", 31), ("O", 32), ("M", 33),
            ]
            assertions = [
                code for code, idx in assertion_map
                if len(row) > idx and row[idx] == "O"
            ]

            controls.append({
                "code": c_code,
                "name": c_name,
                "description": str(row[17] or "").strip() or None,
                "objective": str(row[15] or "").strip() or None,
                "owner_name": str(row[7] or "").strip() or None,
                "risk_code": r_code,
                "is_key_control": (row[18] == "Yes") if len(row) > 18 else True,
                "preventive_detective": pd_val,
                "auto_manual": am_val,
                "activity_approval": len(row) > 19 and row[19] == "O",
                "activity_verification": len(row) > 20 and row[20] == "O",
                "activity_physical": len(row) > 21 and row[21] == "O",
                "activity_master_data": len(row) > 22 and row[22] == "O",
                "activity_reconciliation": len(row) > 23 and row[23] == "O",
                "activity_supervision": len(row) > 24 and row[24] == "O",
                "assertions": assertions,
                "related_accounts": str(row[34] or "").strip() or None,
                "frequency": freq_val,
                "ipe_relevant": ipe_val,
                "related_systems": str(row[37] or "").strip() or None,
                "euc_description": str(row[38] or "").strip() or None,
            })
        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")

    return _ParsedRCM(
        processes=processes,
        sub_processes=sub_processes,
        risks=risks,
        controls=controls,
        errors=errors,
        warnings=warnings,
    )


def _build_not_found_response(sheet_names: list[str], current_range: int) -> JSONResponse:
    """헤더 미발견 시 단계적 확장 또는 최종 오류 응답."""
    next_stage = {15: 30, 30: 130}

    if current_range in next_stage:
        next_range = next_stage[current_range]
        return JSONResponse(
            status_code=200,
            content={
                "status": "needs_expansion",
                "message": (
                    f"1~{current_range}행에서 RCM 헤더를 찾지 못했습니다. "
                    f"{current_range + 1}~{next_range}행까지 확장 검색할까요?"
                ),
                "current_range": current_range,
                "next_range": next_range,
                "expand_param": f"?expand_to={next_range}",
                "sheets_checked": sheet_names,
            },
        )

    return JSONResponse(
        status_code=422,
        content={
            "status": "header_not_found",
            "error": "RCM 헤더가 있는 시트를 찾을 수 없습니다.",
            "checked_sheets": sheet_names,
            "checked_rows": f"1~{current_range}",
            "required_headers": {
                "process_code": "프로세스번호 / 프로세스ID / Process ID / Process No 등",
                "control_code": "통제활동번호 / 통제번호 / Control ID / Control No 등",
                "control_name": "통제활동이름 / 통제명 / Control Name 등",
            },
            "suggestion": (
                "헤더가 다른 이름이거나 130행 이후에 있다면 관리자에게 문의. "
                "동의어 사전 확장이 필요합니다."
            ),
        },
    )


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    mode: str = Form(default="preview"),
    expand_to: int = 15,
    user: CurrentUser = None,
    db: Session = Depends(get_db),
) -> dict:
    """RCM Excel 업로드. 시트명 무관, 헤더 자동 인식.

    mode=preview: 파싱 결과 반환. mode=commit: DB 저장.
    expand_to: 헤더 탐색 최대 행 (1차=15, 2차=30, 3차=130).
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 허용됩니다")
    if mode not in ("preview", "commit"):
        raise HTTPException(status_code=400, detail="mode는 'preview' 또는 'commit'이어야 합니다")

    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 파일 열기 오류: {e}")

    found = find_rcm_sheet(wb, max_row=expand_to)
    if found is None:
        sheet_names = wb.sheetnames
        wb.close()
        return _build_not_found_response(sheet_names, expand_to)

    sheet_name, header_row, mapping = found
    ws = wb[sheet_name]
    try:
        parsed = _parse_rcm_sheet(ws, header_row, mapping)
    except Exception as e:
        wb.close()
        raise HTTPException(status_code=400, detail=f"Excel 파싱 오류: {e}")
    wb.close()

    summary = {
        "total_rows": len(parsed.controls) + len(parsed.errors),
        "valid_rows": len(parsed.controls),
        "errors": parsed.errors,
        "warnings": parsed.warnings,
    }

    if mode == "preview":
        return {
            "summary": summary,
            "preview": parsed.controls[:20],
        }

    # mode == "commit"
    process_id_map: dict = {}
    sp_id_map: dict = {}
    risk_id_map: dict = {}

    for p_code, p_name in parsed.processes.items():
        existing = db.query(Process).filter(Process.code == p_code).first()
        if existing:
            process_id_map[p_code] = existing.id
        else:
            obj = Process(code=p_code, name=p_name or p_code)
            db.add(obj)
            db.flush()
            process_id_map[p_code] = obj.id

    for sp_code, sp_info in parsed.sub_processes.items():
        p_id = process_id_map.get(sp_info["process_code"])
        if not p_id:
            continue
        existing = db.query(SubProcess).filter(SubProcess.code == sp_code).first()
        if existing:
            sp_id_map[sp_code] = existing.id
        else:
            obj = SubProcess(code=sp_code, name=sp_info["name"] or sp_code, process_id=p_id)
            db.add(obj)
            db.flush()
            sp_id_map[sp_code] = obj.id

    for r_code, r_info in parsed.risks.items():
        sp_id = sp_id_map.get(r_info["sub_process_code"])
        if not sp_id:
            continue
        existing = db.query(Risk).filter(Risk.code == r_code).first()
        if existing:
            risk_id_map[r_code] = existing.id
        else:
            obj = Risk(
                code=r_code,
                description=r_info["description"],
                assessment_level=r_info["assessment_level"],
                sub_process_id=sp_id,
            )
            db.add(obj)
            db.flush()
            risk_id_map[r_code] = obj.id

    # 7가지 assertion 코드 캐시
    rc_cache: dict = {}
    for code, name in [("E", "Existence"), ("C", "Completeness"), ("R", "Rights & Obligations"),
                        ("V", "Valuation"), ("P", "Presentation"), ("O", "Occurrence"), ("M", "Measurement")]:
        rc = db.query(RiskCategory).filter(RiskCategory.code == code).first()
        if not rc:
            rc = RiskCategory(code=code, name=name)
            db.add(rc)
            db.flush()
        rc_cache[code] = rc.id

    created = {"processes": 0, "sub_processes": 0, "risks": 0, "controls": 0, "assertions": 0}
    created["processes"] = len([v for v in process_id_map.values()])
    created["sub_processes"] = len([v for v in sp_id_map.values()])
    created["risks"] = len([v for v in risk_id_map.values()])

    for c_data in parsed.controls:
        r_id = risk_id_map.get(c_data["risk_code"])
        if not r_id:
            continue
        existing = db.query(Control).filter(Control.code == c_data["code"]).first()
        if existing:
            continue
        ctrl = Control(
            code=c_data["code"],
            name=c_data["name"],
            description=c_data.get("description"),
            objective=c_data.get("objective"),
            owner_name=c_data.get("owner_name"),
            risk_id=r_id,
            is_key_control=c_data.get("is_key_control", True),
            preventive_detective=c_data.get("preventive_detective", "P"),
            auto_manual=c_data.get("auto_manual", "M"),
            activity_approval=c_data.get("activity_approval", False),
            activity_verification=c_data.get("activity_verification", False),
            activity_physical=c_data.get("activity_physical", False),
            activity_master_data=c_data.get("activity_master_data", False),
            activity_reconciliation=c_data.get("activity_reconciliation", False),
            activity_supervision=c_data.get("activity_supervision", False),
            related_accounts=c_data.get("related_accounts"),
            frequency=c_data.get("frequency", "A"),
            ipe_relevant=c_data.get("ipe_relevant", "N/A"),
            related_systems=c_data.get("related_systems"),
            euc_description=c_data.get("euc_description"),
        )
        db.add(ctrl)
        db.flush()
        created["controls"] += 1

        for a_code in c_data.get("assertions", []):
            rc_id = rc_cache.get(a_code)
            if rc_id:
                db.add(ControlAssertion(control_id=ctrl.id, risk_category_id=rc_id))
                created["assertions"] += 1

    db.commit()
    return {"summary": summary, "created": created}


# ── 위험 매트릭스 ──────────────────────────────────────────

@router.get("/matrix")
def get_matrix(process_code: str | None = None, user: CurrentUser = None, db: Session = Depends(get_db)) -> dict:
    """Process → SubProcess → Risk → Control 중첩 구조. 별도 서비스 클래스 없음 (ADR-0020)."""
    q = db.query(Process).filter(Process.is_deleted == False)  # noqa: E712
    if process_code:
        q = q.filter(Process.code == process_code)
    processes = q.order_by(Process.code).all()

    matrix = []
    total_sp = total_risk = total_ctrl = 0
    level_dist: dict = {"LR": 0, "MR": 0, "HR": 0, "SR": 0}
    freq_dist: dict = {"O": 0, "D": 0, "W": 0, "M": 0, "Q": 0, "A": 0}

    for p in processes:
        sps = db.query(SubProcess).filter(
            SubProcess.process_id == p.id, SubProcess.is_deleted == False  # noqa: E712
        ).order_by(SubProcess.code).all()
        total_sp += len(sps)

        sp_list = []
        for sp in sps:
            risks = db.query(Risk).filter(
                Risk.sub_process_id == sp.id, Risk.is_deleted == False  # noqa: E712
            ).order_by(Risk.code).all()
            total_risk += len(risks)

            risk_list = []
            for r in risks:
                level_dist[r.assessment_level] = level_dist.get(r.assessment_level, 0) + 1
                ctrls = db.query(Control).filter(
                    Control.risk_id == r.id, Control.is_deleted == False  # noqa: E712
                ).order_by(Control.code).all()
                total_ctrl += len(ctrls)

                ctrl_list = []
                for c in ctrls:
                    freq_dist[c.frequency] = freq_dist.get(c.frequency, 0) + 1
                    assertion_codes = [
                        ca.risk_category.code
                        for ca in db.query(ControlAssertion).filter(
                            ControlAssertion.control_id == c.id,
                            ControlAssertion.is_deleted == False,  # noqa: E712
                        ).all()
                        if ca.risk_category
                    ]
                    ctrl_list.append({
                        "code": c.code,
                        "name": c.name,
                        "frequency": c.frequency,
                        "is_key_control": c.is_key_control,
                        "assertions": assertion_codes,
                    })
                risk_list.append({
                    "code": r.code,
                    "description": r.description,
                    "level": r.assessment_level,
                    "controls": ctrl_list,
                })
            sp_list.append({"code": sp.code, "name": sp.name, "risks": risk_list})
        matrix.append({"process": {"code": p.code, "name": p.name}, "sub_processes": sp_list})

    return {
        "matrix": matrix,
        "summary": {
            "process_count": len(processes),
            "sub_process_count": total_sp,
            "risk_count": total_risk,
            "control_count": total_ctrl,
            "risk_level_distribution": level_dist,
            "frequency_distribution": freq_dist,
        },
    }
