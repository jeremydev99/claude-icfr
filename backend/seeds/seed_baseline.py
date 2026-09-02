"""엑셀 단일 원천 → baseline_* 시드 (ADR-0027, 2-A-2 후속 정식화).

2-A-2 이관 스크립트(`scripts/migrate_rcm_to_baseline.py`)는 "마스터 로컬의 기존 controls"
를 복사하는 임시방편이었다. 표준의 유일 원천은 **Regina 보유 최종 엑셀**이며, 이 스크립트가
모든 환경(마스터·Regina·CI·운영)에서 **동일한 baseline 을 재현**한다. seed 가 정식 경로,
2-A-2 이관은 역할 종료(repo 엔 남기되 실행하지 않는다).

실행 (컨테이너 내부, WORKDIR=/app):
    docker compose exec backend python -m seeds.seed_baseline            # 신규 시드
    docker compose exec backend python -m seeds.seed_baseline --reset    # 재구축(삭제 후 시드)

성격:
- baseline 6테이블만 채운다. **instance 는 생성하지 않는다**(암묵 adopt — resolver 가
  instance 없으면 baseline 을 그대로 반환).
- 기존 `controls` 등 구 스키마 6테이블은 **일절 건드리지 않는다**(전후 행 수로 확인 출력).

안전장치:
- 재실행 안전 — baseline 에 데이터가 있으면 중단·보고. 덮어쓰기 옵션 없음.
- `--reset` 은 **명시적 실행 시에만** baseline_* + 그에 연결된 instance 를 비운다.
  실데이터 삭제 경로이므로 기본 실행은 절대 비우지 않는다.
- 단일 트랜잭션 — 중간 실패 시 전부 롤백(부분 시드 없음).
- 검증 출력 — 계층별 삽입 건수, 엑셀 파싱 수와 대조. 불일치 시 롤백 중단.

파서 재사용 (명세 §2):
- `find_rcm_sheet`(services/excel_parser) 와 `_parse_rcm_sheet`(api/rcm) 는 이미
  **controls 에 결합되지 않은 순수 변환**이다 — dict/list 만 반환하고 DB 를 모른다.
  삽입 로직만 upload-excel 의 commit 분기에 있으므로, 추출·복제 없이 그대로 import 한다.
  (파서 코어를 services 로 물리 이동하는 정리는 2-A-4-3 범위)
- 다만 이 엑셀은 헤더가 6~7행 2단이라 `_parse_rcm_sheet` 가 쓰는 "header_row+1 = 데이터 시작"
  가정이 어긋난다(7행이 2차 헤더 → 즉시 break → 0건). 실제 데이터 시작행을 찾아
  `header_row` 인자에 (데이터시작 - 1) 을 넘겨 보정한다. 파서 자체는 무변경.
  보정 함수 `find_data_start_row` 는 `services/excel_parser` 에 있고 upload-excel 과 공용이다
  (13.9-10-a — 이 보정을 seed 만 하고 upload 는 하지 않아 결함이 생겼다).
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app.api.rcm import _parse_rcm_sheet
from app.core.database import SessionLocal
from app.core.tenant_context import (
    DEFAULT_TENANT_CODE,
    reset_active_tenant,
    set_active_tenant,
)
from app.models.rcm_baseline import (
    BaselineControl,
    BaselineControlAssertion,
    BaselineProcess,
    BaselineRisk,
    BaselineRiskCategory,
    BaselineSubProcess,
)
from app.models.tenant import Tenant
from app.services.excel_parser import find_data_start_row, find_rcm_sheet

# 표준 원천 — repo 내 고정 경로. 모든 환경이 같은 파일을 본다.
EXCEL_PATH = Path(__file__).resolve().parent / "2026_설계평가_RCM_리스트.xlsx"

# 어서션 7종 마스터. upload-excel 의 코드→이름 매핑과 동일하게 유지한다
# (엑셀 헤더의 한글 표기 실재성/완전성/권리와의무/평가/재무제표표시와공시/발생사실/측정 과 1:1).
ASSERTION_MASTER = [
    ("E", "Existence"),
    ("C", "Completeness"),
    ("R", "Rights & Obligations"),
    ("V", "Valuation"),
    ("P", "Presentation"),
    ("O", "Occurrence"),
    ("M", "Measurement"),
]

_BASELINE_MODELS = [
    BaselineProcess, BaselineSubProcess, BaselineRisk,
    BaselineRiskCategory, BaselineControl, BaselineControlAssertion,
]

# --reset 삭제 순서 — 자식 먼저. instance 5테이블이 baseline 을 FK 참조하므로 instance 가 앞선다.
# 구 스키마(processes/controls/...)는 목록에 없다 = 건드리지 않는다.
_RESET_TABLES = [
    "control_assertion_instances",
    "control_instances",
    "risk_instances",
    "sub_process_instances",
    "process_instances",
    "baseline_control_assertions",
    "baseline_controls",
    "baseline_risks",
    "baseline_sub_processes",
    "baseline_processes",
    "baseline_risk_categories",
]

# 구 스키마 불변 확인 대상.
_LEGACY_TABLES = [
    "processes", "sub_processes", "risks", "risk_categories", "controls", "control_assertions",
]

# Control 미러링 필드 (risk_id 는 코드→id 매핑, baseline_version 은 default=1).
_CONTROL_FIELDS = [
    "code", "name", "description",
    "objective", "owner_name",
    "is_key_control", "preventive_detective", "auto_manual",
    "activity_approval", "activity_verification", "activity_physical",
    "activity_master_data", "activity_reconciliation", "activity_supervision",
    "related_accounts", "frequency", "ipe_relevant", "related_systems", "euc_description",
]


def _load_excel():
    """엑셀 → 파싱 결과(_ParsedRCM). 순수 변환만 사용."""
    if not EXCEL_PATH.exists():
        raise SystemExit(
            f"[중단] 원천 엑셀이 없습니다: {EXCEL_PATH}\n"
            "  이미지에 seeds/ 가 포함됐는지 확인하세요(Dockerfile COPY seeds/)."
        )

    wb = load_workbook(EXCEL_PATH, data_only=True)
    try:
        found = find_rcm_sheet(wb)
        if found is None:
            raise SystemExit(f"[중단] RCM 헤더를 찾지 못했습니다: 시트 {wb.sheetnames}")
        sheet_name, header_row, mapping = found
        ws = wb[sheet_name]
        scan = 10
        data_start = find_data_start_row(ws, header_row, mapping["process_code"], max_scan=scan)
        if data_start is None:
            raise SystemExit(
                f"[중단] 헤더행 {header_row} 다음 {scan}행 안에서 데이터 시작행을 찾지 못했습니다."
            )
        print(f"  원천: {EXCEL_PATH.name} / 시트 {sheet_name} / 헤더 {header_row}행 / 데이터 {data_start}행~")
        parsed = _parse_rcm_sheet(ws, data_start - 1, mapping)
    finally:
        wb.close()

    if parsed.errors:
        raise SystemExit(f"[중단] 엑셀 파싱 오류 {len(parsed.errors)}건: {parsed.errors[:10]}")
    for w in parsed.warnings:
        print(f"  [경고] {w}")
    return parsed


def _count_baseline(db) -> dict[str, int]:
    return {m.__tablename__: db.query(m).count() for m in _BASELINE_MODELS}


def _count_tables(db, tables: list[str]) -> dict[str, int]:
    return {t: db.execute(text(f"SELECT count(*) FROM {t}")).scalar() for t in tables}


def _assert_baseline_empty(db) -> None:
    """재실행 안전 — baseline 에 데이터가 하나라도 있으면 중단."""
    existing = _count_baseline(db)
    nonempty = {t: c for t, c in existing.items() if c > 0}
    if nonempty:
        raise SystemExit(
            f"[중단] baseline 테이블에 이미 데이터가 있습니다: {nonempty}\n"
            "  시드는 baseline 이 비어 있을 때만 실행됩니다. 덮어쓰기 옵션은 제공하지 않습니다.\n"
            "  엑셀 기준으로 재구축하려면 --reset 을 명시하세요(실데이터 삭제 경로)."
        )


def _resolve_tenant(db, code: str):
    """대상 테넌트 code → id. 없으면 중단 (ADR-0030 §2.5 — 하드코딩 금지, DB 조회).

    baseline 이 테넌트 소유가 되면서 "어느 회사의 표준을 시드하는가"가 필수 정보가 됐다.
    id 를 코드에 박지 않고 `tenants` 에서 조회한다 — 환경마다 id 가 다를 수 있다.
    """
    tenant = db.query(Tenant).filter(Tenant.code == code).first()
    if tenant is None:
        existing = [t.code for t in db.query(Tenant).all()]
        raise SystemExit(
            f"[중단] 테넌트 코드 '{code}' 를 찾을 수 없습니다. 존재하는 코드: {existing}"
        )
    return tenant.id, tenant.name


def _reset(db) -> None:
    """baseline_* + 그에 연결된 instance 를 비운다. 구 스키마는 대상 아님."""
    before = _count_tables(db, _RESET_TABLES)
    print("  [reset] 삭제 전 행 수:")
    for t in _RESET_TABLES:
        print(f"    {t:32s}{before[t]:6d}")

    # ORM 이 아닌 raw DELETE — instance 는 AuditedBase(tenant 자동 필터) 이므로
    # 전 tenant 의 행을 남김없이 지우려면 ORM 필터를 우회해야 한다.
    for t in _RESET_TABLES:
        db.execute(text(f"DELETE FROM {t}"))
    db.flush()

    after = _count_tables(db, _RESET_TABLES)
    remaining = {t: c for t, c in after.items() if c > 0}
    if remaining:
        raise SystemExit(f"[중단] reset 후에도 남은 행이 있습니다: {remaining}")
    print(f"  [reset] 삭제 완료 — 총 {sum(before.values())}행 제거")


def _seed(db, parsed) -> dict[str, int]:
    """계층 순서로 baseline_* 삽입 + id 매핑. baseline_version 은 default=1."""
    map_process: dict[str, object] = {}
    for code, name in parsed.processes.items():
        obj = BaselineProcess(code=code, name=name or code)
        db.add(obj)
        db.flush()
        map_process[code] = obj.id

    map_sub: dict[str, object] = {}
    for code, info in parsed.sub_processes.items():
        p_id = map_process.get(info["process_code"])
        if p_id is None:
            raise SystemExit(f"[중단] 하위프로세스 {code}: 상위 프로세스 {info['process_code']} 없음")
        obj = BaselineSubProcess(code=code, name=info["name"] or code, process_id=p_id)
        db.add(obj)
        db.flush()
        map_sub[code] = obj.id

    map_risk: dict[str, object] = {}
    for code, info in parsed.risks.items():
        sp_id = map_sub.get(info["sub_process_code"])
        if sp_id is None:
            raise SystemExit(f"[중단] 위험 {code}: 상위 하위프로세스 {info['sub_process_code']} 없음")
        obj = BaselineRisk(
            code=code,
            description=info["description"],
            assessment_level=info["assessment_level"],
            sub_process_id=sp_id,
        )
        db.add(obj)
        db.flush()
        map_risk[code] = obj.id

    # 어서션 7종 마스터 먼저 (통제마다 반복 등장하므로 유니크 삽입 후 연결).
    map_rc: dict[str, object] = {}
    for code, name in ASSERTION_MASTER:
        obj = BaselineRiskCategory(code=code, name=name)
        db.add(obj)
        db.flush()
        map_rc[code] = obj.id

    assertion_links = 0
    for c in parsed.controls:
        r_id = map_risk.get(c["risk_code"])
        if r_id is None:
            raise SystemExit(f"[중단] 통제 {c['code']}: 상위 위험 {c['risk_code']} 없음")
        ctrl = BaselineControl(risk_id=r_id, **{f: c[f] for f in _CONTROL_FIELDS})
        db.add(ctrl)
        db.flush()

        for a_code in c["assertions"]:
            rc_id = map_rc.get(a_code)
            if rc_id is None:
                raise SystemExit(f"[중단] 통제 {c['code']}: 미지의 어서션 코드 {a_code!r}")
            db.add(BaselineControlAssertion(
                baseline_control_id=ctrl.id, baseline_risk_category_id=rc_id,
            ))
            assertion_links += 1
    db.flush()

    return {
        "processes": len(parsed.processes),
        "sub_processes": len(parsed.sub_processes),
        "risks": len(parsed.risks),
        "risk_categories": len(ASSERTION_MASTER),
        "controls": len(parsed.controls),
        "control_assertions": assertion_links,
    }


def seed(reset: bool = False, tenant_code: str = DEFAULT_TENANT_CODE) -> None:
    parsed = _load_excel()
    print(
        f"  파싱: 프로세스 {len(parsed.processes)} / 하위프로세스 {len(parsed.sub_processes)} / "
        f"위험 {len(parsed.risks)} / 통제 {len(parsed.controls)}"
    )

    db = SessionLocal()
    tok = None
    try:
        tenant_id, tenant_name = _resolve_tenant(db, tenant_code)
        print(f"  대상 테넌트: {tenant_name} ({tenant_code}) / {tenant_id}")
        # 활성 tenant 설정 → before_flush 가 baseline 행에 tenant_id 를 자동 stamp 한다
        # (ADR-0025 자동 격리. 수동 지정 금지 — ADR-0030 전환으로 baseline 도 대상이 됐다).
        tok = set_active_tenant(tenant_id)

        legacy_before = _count_tables(db, _LEGACY_TABLES)

        if reset:
            _reset(db)
        else:
            _assert_baseline_empty(db)

        expected = _seed(db, parsed)

        # ── 검증 (엑셀 파싱 수 vs baseline 실제 행 수) ──
        pairs = [
            ("processes", expected["processes"], BaselineProcess),
            ("sub_processes", expected["sub_processes"], BaselineSubProcess),
            ("risks", expected["risks"], BaselineRisk),
            ("risk_categories", expected["risk_categories"], BaselineRiskCategory),
            ("controls", expected["controls"], BaselineControl),
            ("control_assertions", expected["control_assertions"], BaselineControlAssertion),
        ]
        mismatch = []
        print("\n  검증 (엑셀 기준 -> baseline):")
        for name, src_n, model in pairs:
            dst_n = db.query(model).count()
            marker = "OK" if src_n == dst_n else "MISMATCH"
            print(f"    {name:20s}{src_n:5d} -> {model.__tablename__:32s}{dst_n:5d}  [{marker}]")
            if src_n != dst_n:
                mismatch.append((name, src_n, dst_n))
        if mismatch:
            db.rollback()
            raise SystemExit(f"[중단] 파싱/삽입 수 불일치 — 전체 롤백: {mismatch}")

        # ── instance 미생성 확인 ──
        inst = _count_tables(db, _RESET_TABLES[:5])
        if any(inst.values()):
            db.rollback()
            raise SystemExit(f"[중단] instance 가 생성됐습니다(암묵 adopt 위반) — 롤백: {inst}")
        print("    instance 5테이블         0건 (암묵 adopt)  [OK]")

        # ── 구 스키마 불변 확인 ──
        legacy_after = _count_tables(db, _LEGACY_TABLES)
        if legacy_before != legacy_after:
            db.rollback()
            raise SystemExit(f"[중단] 구 스키마가 변경됐습니다 — 롤백: {legacy_before} -> {legacy_after}")
        print(f"    구 스키마 불변          {legacy_after}  [OK]")

        db.commit()
        print("\n✓ 시드 완료 — 단일 트랜잭션 커밋.")
    except SystemExit:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        print("[에러] 예외 발생 — 전체 롤백했습니다.")
        raise
    finally:
        if tok is not None:
            reset_active_tenant(tok)
        db.close()


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="엑셀 단일 원천 → baseline_* 시드")
    ap.add_argument(
        "--reset",
        action="store_true",
        help="기존 baseline_* 와 연결된 instance 를 비우고 재삽입 (실데이터 삭제 경로)",
    )
    ap.add_argument(
        "--tenant",
        default=DEFAULT_TENANT_CODE,
        help=f"baseline 을 귀속시킬 테넌트 code (기본 {DEFAULT_TENANT_CODE}). id 는 DB 에서 조회한다",
    )
    args = ap.parse_args()
    seed(reset=args.reset, tenant_code=args.tenant)
