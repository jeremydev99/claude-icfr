"""RCM 모듈 통합 테스트 — Phase 1 풀 확장."""
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook


def _token(client: TestClient) -> str:
    resp = client.post("/api/auth/login", data={"username": "admin@acme.example", "password": "admin123"})
    assert resp.status_code == 200
    return resp.json()["access_token"]


def _headers(client: TestClient) -> dict:
    return {"Authorization": f"Bearer {_token(client)}"}


def _make_test_excel(
    p_code="TP", p_name="테스트프로세스",
    sp_code="TP-010", sp_name="테스트하위",
    r_code="TP-010-10",
    c_code="TP-010-10-10", c_name="테스트통제",
) -> bytes:
    """최소 RCM Excel 파일 생성 (테스트용). 헤더 7행, 데이터 8행."""
    wb = Workbook()
    ws = wb.active
    ws.title = "RCM"
    for _ in range(6):
        ws.append([None] * 45)
    # 헤더 행 (row 7) — 필수 3개 컬럼 동의어 사전 매칭
    hdr = [None] * 45
    hdr[1] = "프로세스번호"
    hdr[6] = "통제활동번호"
    hdr[16] = "통제활동이름"
    ws.append(hdr)
    # 데이터 행 (row 8)
    row = [None] * 45
    row[1] = p_code    # B
    row[2] = p_name    # C
    row[3] = sp_code   # D
    row[4] = sp_name   # E
    row[5] = r_code    # F
    row[6] = c_code    # G
    row[7] = "관리자"   # H
    row[8] = "테스트위험"  # I
    row[14] = "LR"     # O
    row[15] = "통제목적"  # P
    row[16] = c_name   # Q
    row[17] = "통제설명"  # R
    row[18] = "Yes"    # S
    row[19] = "O"      # T
    row[25] = "P"      # Z
    row[26] = "M"      # AA
    row[27] = "O"      # AB — assertion E
    row[28] = "O"      # AC — assertion C
    row[35] = "M"      # AJ
    row[36] = "N/A"    # AK
    ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_excel_with_headers(
    sheet_name: str = "RCM",
    header_row: int = 7,
    pc_header: str = "프로세스번호",
    cc_header: str = "통제활동번호",
    cn_header: str = "통제활동이름",
    p_code: str = "TP", p_name: str = "테스트프로세스",
    sp_code: str = "TP-010", sp_name: str = "테스트하위",
    r_code: str = "TP-010-10",
    c_code: str = "TP-010-10-10", c_name: str = "테스트통제",
) -> bytes:
    """헤더 위치·이름·시트명을 자유롭게 지정할 수 있는 테스트용 Excel 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for _ in range(header_row - 1):
        ws.append([None] * 45)
    hdr = [None] * 45
    hdr[1] = pc_header
    hdr[6] = cc_header
    hdr[16] = cn_header
    ws.append(hdr)
    row = [None] * 45
    row[1] = p_code
    row[2] = p_name
    row[3] = sp_code
    row[4] = sp_name
    row[5] = r_code
    row[6] = c_code
    row[7] = "관리자"
    row[8] = "테스트위험"
    row[14] = "LR"
    row[15] = "통제목적"
    row[16] = c_name
    row[17] = "통제설명"
    row[18] = "Yes"
    row[25] = "P"
    row[26] = "M"
    row[27] = "O"
    row[35] = "M"
    row[36] = "N/A"
    ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Process CRUD ──────────────────────────────────────────

def test_process_crud(client: TestClient) -> None:
    h = _headers(client)
    resp = client.post("/api/rcm/processes", json={"code": "P-TEST", "name": "테스트 프로세스"}, headers=h)
    assert resp.status_code == 201
    pid = resp.json()["id"]

    resp = client.get(f"/api/rcm/processes/{pid}", headers=h)
    assert resp.status_code == 200

    resp = client.get("/api/rcm/processes", headers=h)
    assert resp.json()["total"] >= 1

    resp = client.patch(f"/api/rcm/processes/{pid}", json={"name": "수정된 프로세스"}, headers=h)
    assert resp.status_code == 200
    assert resp.json()["name"] == "수정된 프로세스"

    resp = client.delete(f"/api/rcm/processes/{pid}", headers=h)
    assert resp.status_code == 204

    resp = client.get(f"/api/rcm/processes/{pid}", headers=h)
    assert resp.status_code == 404


# ── SubProcess CRUD ───────────────────────────────────────

def test_subprocess_crud(client: TestClient) -> None:
    h = _headers(client)
    p = client.post("/api/rcm/processes", json={"code": "P-SP-TEST", "name": "SP 테스트 프로세스"}, headers=h)
    pid = p.json()["id"]

    resp = client.post("/api/rcm/sub-processes", json={"code": "SP-010", "name": "하위프로세스", "process_id": pid}, headers=h)
    assert resp.status_code == 201
    sp_id = resp.json()["id"]
    assert resp.json()["code"] == "SP-010"

    resp = client.get(f"/api/rcm/sub-processes/{sp_id}", headers=h)
    assert resp.status_code == 200

    resp = client.get("/api/rcm/sub-processes", headers=h)
    assert resp.json()["total"] >= 1

    resp = client.patch(f"/api/rcm/sub-processes/{sp_id}", json={"name": "수정된 하위프로세스"}, headers=h)
    assert resp.status_code == 200

    resp = client.delete(f"/api/rcm/sub-processes/{sp_id}", headers=h)
    assert resp.status_code == 204


# ── Risk CRUD ─────────────────────────────────────────────

def test_risk_crud(client: TestClient) -> None:
    h = _headers(client)
    p = client.post("/api/rcm/processes", json={"code": "P-RISK-TEST", "name": "Risk 테스트"}, headers=h)
    sp = client.post("/api/rcm/sub-processes", json={"code": "SP-RISK-010", "name": "SP", "process_id": p.json()["id"]}, headers=h)

    resp = client.post("/api/rcm/risks", json={
        "code": "RISK-010-10",
        "description": "테스트 위험",
        "assessment_level": "MR",
        "sub_process_id": sp.json()["id"],
    }, headers=h)
    assert resp.status_code == 201
    risk_id = resp.json()["id"]
    assert resp.json()["assessment_level"] == "MR"

    resp = client.get(f"/api/rcm/risks/{risk_id}", headers=h)
    assert resp.status_code == 200

    resp = client.patch(f"/api/rcm/risks/{risk_id}", json={"assessment_level": "HR"}, headers=h)
    assert resp.json()["assessment_level"] == "HR"

    resp = client.delete(f"/api/rcm/risks/{risk_id}", headers=h)
    assert resp.status_code == 204


# ── Control CRUD (새 구조) ─────────────────────────────────

def test_control_extended_crud(client: TestClient) -> None:
    h = _headers(client)
    p = client.post("/api/rcm/processes", json={"code": "P-CTL-EXT", "name": "통제 테스트"}, headers=h)
    sp = client.post("/api/rcm/sub-processes", json={"code": "SP-CTL-010", "name": "SP", "process_id": p.json()["id"]}, headers=h)
    r = client.post("/api/rcm/risks", json={
        "code": "CTL-RISK-010", "description": "위험", "assessment_level": "LR", "sub_process_id": sp.json()["id"]
    }, headers=h)
    risk_id = r.json()["id"]

    resp = client.post("/api/rcm/controls", json={
        "code": "CTL-EXT-001",
        "name": "확장 통제",
        "risk_id": risk_id,
        "frequency": "M",
        "is_key_control": True,
        "preventive_detective": "P",
        "auto_manual": "M",
        "activity_approval": True,
        "ipe_relevant": "N/A",
    }, headers=h)
    assert resp.status_code == 201
    cid = resp.json()["id"]
    assert resp.json()["frequency"] == "M"
    assert resp.json()["activity_approval"] is True

    resp = client.get(f"/api/rcm/controls/{cid}", headers=h)
    assert resp.status_code == 200

    resp = client.patch(f"/api/rcm/controls/{cid}", json={"frequency": "Q", "auto_manual": "A"}, headers=h)
    assert resp.json()["frequency"] == "Q"
    assert resp.json()["auto_manual"] == "A"

    resp = client.get("/api/rcm/controls", headers=h)
    assert resp.json()["total"] >= 1

    resp = client.delete(f"/api/rcm/controls/{cid}", headers=h)
    assert resp.status_code == 204


# ── Excel 업로드 ──────────────────────────────────────────

def test_excel_upload_preview(client: TestClient) -> None:
    h = _headers(client)
    excel = _make_test_excel()
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test_rcm.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert "summary" in data
    assert data["summary"]["valid_rows"] == 1
    assert len(data["preview"]) == 1
    assert data["preview"][0]["code"] == "TP-010-10-10"


def test_excel_upload_commit(client: TestClient) -> None:
    h = _headers(client)
    excel = _make_test_excel(
        p_code="EC-P", sp_code="EC-P-010", r_code="EC-P-010-10",
        c_code="EC-P-010-10-10", c_name="Excel Commit 통제"
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test_rcm.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "commit"},
        headers=h,
    )
    assert resp.status_code == 200
    created = resp.json()["created"]
    assert created["controls"] == 1
    assert created["assertions"] >= 1

    # 실제로 DB에 저장됐는지 확인
    resp = client.get("/api/rcm/controls", headers=h, params={"limit": 200})
    codes = [c["code"] for c in resp.json()["items"]]
    assert "EC-P-010-10-10" in codes


# ── 검색 API ──────────────────────────────────────────────

def test_search_text(client: TestClient) -> None:
    h = _headers(client)
    # 검색 전 통제 생성
    p = client.post("/api/rcm/processes", json={"code": "SRCH-P", "name": "검색 프로세스"}, headers=h)
    sp = client.post("/api/rcm/sub-processes", json={"code": "SRCH-SP", "name": "검색 SP", "process_id": p.json()["id"]}, headers=h)
    r = client.post("/api/rcm/risks", json={"code": "SRCH-R", "description": "위험", "assessment_level": "LR", "sub_process_id": sp.json()["id"]}, headers=h)
    client.post("/api/rcm/controls", json={"code": "SRCH-CTL-001", "name": "검색용통제명칭", "risk_id": r.json()["id"], "frequency": "M"}, headers=h)

    resp = client.get("/api/rcm/controls/search", params={"q": "검색용통제"}, headers=h)
    assert resp.status_code == 200
    codes = [c["code"] for c in resp.json()["items"]]
    assert "SRCH-CTL-001" in codes


def test_search_filter_combination(client: TestClient) -> None:
    h = _headers(client)
    resp = client.get("/api/rcm/controls/search", params={"frequency": "M", "is_key_control": "true"}, headers=h)
    assert resp.status_code == 200
    assert "total" in resp.json()


# ── 매트릭스 API ──────────────────────────────────────────

def test_matrix_endpoint(client: TestClient) -> None:
    h = _headers(client)
    resp = client.get("/api/rcm/matrix", headers=h)
    assert resp.status_code == 200
    data = resp.json()
    assert "matrix" in data
    assert "summary" in data
    assert "process_count" in data["summary"]


# ── 벌크 작업 ─────────────────────────────────────────────

def test_bulk_delete(client: TestClient) -> None:
    h = _headers(client)
    p = client.post("/api/rcm/processes", json={"code": "BLK-P", "name": "벌크 삭제 테스트"}, headers=h)
    sp = client.post("/api/rcm/sub-processes", json={"code": "BLK-SP", "name": "SP", "process_id": p.json()["id"]}, headers=h)
    r = client.post("/api/rcm/risks", json={"code": "BLK-R", "description": "위험", "assessment_level": "LR", "sub_process_id": sp.json()["id"]}, headers=h)
    c1 = client.post("/api/rcm/controls", json={"code": "BLK-C-001", "name": "벌크1", "risk_id": r.json()["id"], "frequency": "A"}, headers=h)
    c2 = client.post("/api/rcm/controls", json={"code": "BLK-C-002", "name": "벌크2", "risk_id": r.json()["id"], "frequency": "A"}, headers=h)

    resp = client.post("/api/rcm/controls/bulk-delete", json={"control_ids": [c1.json()["id"], c2.json()["id"]]}, headers=h)
    assert resp.status_code == 200
    assert resp.json()["deleted_count"] == 2


def test_clear_all(client: TestClient) -> None:
    h = _headers(client)
    # clear-all 이후 controls가 0이어야 함
    resp = client.post("/api/rcm/controls/clear-all", json={"confirm": "DELETE_ALL_RCM_DATA"}, headers=h)
    assert resp.status_code == 200
    deleted = resp.json()["deleted"]
    assert "controls" in deleted

    resp = client.get("/api/rcm/controls", headers=h)
    assert resp.json()["total"] == 0


# ── Excel 헤더 자동 인식 테스트 ──────────────────────────────

def test_upload_excel_sheet_name_arbitrary(client: TestClient) -> None:
    """시트명이 'RCM'이 아니어도 헤더 매칭 시 인식."""
    h = _headers(client)
    excel = _make_excel_with_headers(
        sheet_name="통제매트릭스",
        c_code="SN-001", c_name="시트명무관통제",
        p_code="SN-P", sp_code="SN-SP", r_code="SN-R",
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["summary"]["valid_rows"] == 1
    assert data["preview"][0]["code"] == "SN-001"


def test_upload_excel_header_row_position_varies(client: TestClient) -> None:
    """헤더가 8행이 아닌 5행에 있어도 인식."""
    h = _headers(client)
    excel = _make_excel_with_headers(
        header_row=5,
        c_code="HR-001", c_name="헤더위치5행통제",
        p_code="HR-P", sp_code="HR-SP", r_code="HR-R",
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    assert resp.json()["summary"]["valid_rows"] == 1


def test_upload_excel_korean_synonyms(client: TestClient) -> None:
    """헤더 이름이 동의어 ('통제코드', '프로세스코드') 인 경우 인식."""
    h = _headers(client)
    excel = _make_excel_with_headers(
        pc_header="프로세스코드",
        cc_header="통제코드",
        cn_header="통제명",
        c_code="KS-001", c_name="한글동의어통제",
        p_code="KS-P", sp_code="KS-SP", r_code="KS-R",
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    assert resp.json()["summary"]["valid_rows"] == 1


def test_upload_excel_english_synonyms(client: TestClient) -> None:
    """헤더가 영문 (Control ID 등) 인 경우 인식."""
    h = _headers(client)
    excel = _make_excel_with_headers(
        pc_header="Process Code",
        cc_header="Control ID",
        cn_header="Control Name",
        c_code="EN-001", c_name="English Header Control",
        p_code="EN-P", sp_code="EN-SP", r_code="EN-R",
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    assert resp.json()["summary"]["valid_rows"] == 1


def test_upload_excel_normalization(client: TestClient) -> None:
    """공백·대소문자 차이 ('Process  ID', 'control name') 도 인식."""
    h = _headers(client)
    excel = _make_excel_with_headers(
        pc_header="Process  ID",   # 공백 2개
        cc_header="control id",    # 소문자
        cn_header="Control  Name", # 공백 2개
        c_code="NM-001", c_name="정규화통제",
        p_code="NM-P", sp_code="NM-SP", r_code="NM-R",
    )
    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    assert resp.json()["summary"]["valid_rows"] == 1


def test_upload_excel_no_match_returns_expansion(client: TestClient) -> None:
    """헤더 못 찾으면 needs_expansion 응답 (1차 시도, expand_to=15)."""
    h = _headers(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for _ in range(20):
        ws.append(["무관한데이터", "abc", None])
    buf = BytesIO()
    wb.save(buf)
    excel = buf.getvalue()

    resp = client.post(
        "/api/rcm/upload-excel",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "needs_expansion"
    assert data["next_range"] == 30


def test_upload_excel_no_match_after_130_returns_error(client: TestClient) -> None:
    """130행까지 시도해도 못 찾으면 422 + 가이드."""
    h = _headers(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for _ in range(5):
        ws.append(["무관한데이터"])
    buf = BytesIO()
    wb.save(buf)
    excel = buf.getvalue()

    resp = client.post(
        "/api/rcm/upload-excel?expand_to=130",
        files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": "preview"},
        headers=h,
    )
    assert resp.status_code == 422
    data = resp.json()
    assert data["status"] == "header_not_found"
    assert "required_headers" in data
