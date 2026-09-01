"""멀티헤더 Excel 파싱 회귀 (13.6 / 13.9-10-a).

이 결함이 살아남은 이유는 **Excel 테스트 8건이 전부 단일 헤더 픽스처를 썼기 때문**이다.
그래서 여기서는 두 가지를 함께 둔다.

- 합성 픽스처(`_make_multiheader_excel`) — 구조를 코드로 고정한다. 원천 파일이 교체돼도
  "2단 헤더를 읽는다"는 계약은 이쪽이 지킨다.
- 실파일(`seeds/2026_설계평가_RCM_리스트.xlsx`) — 합성 픽스처가 실제 구조와 어긋날 수
  있으므로 실증도 함께 둔다. 파일이 없는 환경에서는 skip(test_seed_baseline 선례).

둘 중 하나만으로는 부족하다. 합성만 두면 "재현했다고 믿은 구조"만 검증하고,
실파일만 두면 파일이 바뀌는 순간 무엇을 지키던 테스트였는지 알 수 없게 된다.
"""
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

XLSX = Path(__file__).resolve().parents[1] / "seeds" / "2026_설계평가_RCM_리스트.xlsx"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _headers(client: TestClient) -> dict:
    resp = client.post("/api/auth/login", data={"username": "admin@acme.example", "password": "admin123"})
    assert resp.status_code == 200
    return {"Authorization": "Bearer " + resp.json()["access_token"]}


def _make_multiheader_excel(rows: int = 2, second_header: bool = True) -> bytes:
    """실제 원천 파일 구조를 재현한다 — 1~2행 메타 / 3~4행 공백 / 5행 설명 밴드 /
    6행 주헤더 / 7행 2차 헤더(오른쪽 열에만 값) / 8행부터 데이터.

    2차 헤더 행의 필수 열(프로세스번호)이 비어 있는 것이 핵심이다. 보정 없이 파싱하면
    바로 이 행에서 멈춘다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "RCM"

    meta1 = [None] * 45
    meta1[0], meta1[3] = "Company Name:", "테스트회사"
    ws.append(meta1)
    meta2 = [None] * 45
    meta2[0], meta2[3] = "Update:", "2026-09-01"
    ws.append(meta2)
    ws.append([None] * 45)                      # 3행 공백
    ws.append([None] * 45)                      # 4행 공백

    band = [None] * 45                          # 5행 상단 설명 밴드
    band[9], band[25] = "계정과목 등의 금액을 결정", "예방(P) VS 적발(D)"
    ws.append(band)

    hdr = [None] * 45                           # 6행 주헤더
    hdr[1], hdr[6], hdr[16] = "프로세스번호", "통제활동번호", "통제활동이름"
    ws.append(hdr)

    if second_header:                           # 7행 2차 헤더 — 필수 열은 비어 있다
        sub = [None] * 45
        sub[40], sub[41] = "테스트방법", "모집단"
        ws.append(sub)

    for i in range(1, rows + 1):                # 8행~ 데이터
        row = [None] * 45
        row[1], row[2] = "MH", "멀티헤더프로세스"
        row[3], row[4] = "MH-010", "멀티헤더하위"
        row[5] = f"MH-010-{i}0"
        row[6] = f"MH-010-{i}0-10"
        row[7], row[8] = "관리자", "멀티헤더위험"
        row[14], row[15] = "LR", "통제목적"
        row[16], row[17] = f"멀티헤더통제{i}", "통제활동설명"
        row[18] = "Yes"
        row[25], row[26] = "P", "M"
        row[27] = "O"                           # 어서션 E
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client: TestClient, h: dict, content: bytes, name: str = "multiheader.xlsx"):
    return client.post(
        "/api/rcm/upload-excel",
        files={"file": (name, content, _XLSX_MIME)},
        data={"mode": "preview"},
        headers=h,
    )


# ── 1. 2차 헤더가 있어도 데이터가 파싱될 것 ────────────────

def test_multiheader_sheet_parses_data_rows(client: TestClient) -> None:
    """2단 헤더 구조에서 데이터가 정상 파싱된다 (보정 전에는 0건이었다)."""
    h = _headers(client)
    resp = _upload(client, h, _make_multiheader_excel(rows=2))
    assert resp.status_code == 200, resp.text
    summary = resp.json()["summary"]
    assert summary["valid_rows"] == 2
    assert summary["errors"] == []
    codes = [c["code"] for c in resp.json()["preview"]]
    assert codes == ["MH-010-10-10", "MH-010-20-10"]


def test_multiheader_and_single_header_agree(client: TestClient) -> None:
    """같은 데이터면 2차 헤더 유무와 무관하게 같은 결과 — 보정이 단일 헤더를 해치지 않는다."""
    h = _headers(client)
    multi = _upload(client, h, _make_multiheader_excel(rows=2, second_header=True)).json()
    single = _upload(client, h, _make_multiheader_excel(rows=2, second_header=False)).json()
    assert multi["summary"]["valid_rows"] == single["summary"]["valid_rows"] == 2
    assert [c["code"] for c in multi["preview"]] == [c["code"] for c in single["preview"]]


# ── 2. 0건을 성공으로 반환하지 않을 것 ──────────────────────

def test_no_data_rows_is_not_success(client: TestClient) -> None:
    """헤더는 인식했으나 데이터가 없으면 422 — 200/0건으로 새지 않는다.

    0건을 성공으로 돌려주면 93건짜리 RCM 을 올리고도 "업로드 성공"이 뜬다.
    """
    h = _headers(client)
    resp = _upload(client, h, _make_multiheader_excel(rows=0))
    assert resp.status_code == 422, resp.text
    body = resp.json()
    assert body["status"] == "no_data_rows"
    assert body["header_row"] == 6
    assert body["reason"]        # 다음 행동을 짐작할 수 있어야 한다
    assert body["suggestion"]


def test_no_data_rows_uses_existing_error_shape(client: TestClient) -> None:
    """실패 형태가 기존 규약(header_not_found)과 같다 — 프론트가 따로 분기하지 않도록."""
    h = _headers(client)
    empty = Workbook()
    empty.active.append(["헤더없음"])
    buf = BytesIO()
    empty.save(buf)

    # 헤더 미발견의 **최종** 실패 형태는 확장 한계(130행)까지 갔을 때 나온다.
    # 그 전 단계는 200 + needs_expansion(확장 제안)이라 비교 대상이 아니다.
    no_header = client.post(
        "/api/rcm/upload-excel?expand_to=130",
        files={"file": ("noheader.xlsx", buf.getvalue(), _XLSX_MIME)},
        data={"mode": "preview"}, headers=h,
    )
    no_rows = _upload(client, h, _make_multiheader_excel(rows=0))
    assert no_header.status_code == no_rows.status_code == 422
    for body in (no_header.json(), no_rows.json()):
        assert "status" in body and "error" in body and "suggestion" in body


# ── 3. 단일 헤더 회귀 가드 ────────────────────────────────

def test_single_header_still_parses(client: TestClient) -> None:
    """단일 헤더(헤더 다음 행이 곧 데이터)는 기존대로 동작한다."""
    h = _headers(client)
    wb = Workbook()
    ws = wb.active
    ws.title = "RCM"
    hdr = [None] * 45
    hdr[1], hdr[6], hdr[16] = "프로세스번호", "통제활동번호", "통제활동이름"
    ws.append(hdr)                              # 1행 헤더
    row = [None] * 45
    row[1], row[2] = "SH", "단일헤더프로세스"
    row[3], row[4] = "SH-010", "단일헤더하위"
    row[5], row[6] = "SH-010-10", "SH-010-10-10"
    row[7], row[8] = "관리자", "단일헤더위험"
    row[14], row[15], row[16] = "LR", "통제목적", "단일헤더통제"
    row[25], row[26], row[27] = "P", "M", "O"
    ws.append(row)                              # 2행 데이터
    buf = BytesIO()
    wb.save(buf)

    resp = _upload(client, h, buf.getvalue(), name="single.xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.json()["summary"]["valid_rows"] == 1
    assert resp.json()["preview"][0]["code"] == "SH-010-10-10"


# ── 4. 실파일 실증 ────────────────────────────────────────

@pytest.mark.skipif(not XLSX.exists(), reason=f"원천 엑셀 없음: {XLSX}")
def test_real_rcm_file_parses_all_controls(client: TestClient) -> None:
    """원천 엑셀을 엔드포인트로 처리하면 93건 — 0건이 재현되던 바로 그 경로다.

    건수를 하드코딩하지 않고 seed 파서 결과와 대조한다. 두 경로가 같은 파일을 다르게
    읽던 것이 결함의 본질이므로, 검증도 "두 경로가 일치하는가"로 두는 것이 맞다.
    """
    from seeds.seed_baseline import _load_excel

    expected = len(_load_excel().controls)
    assert expected == 93, f"원천 파일이 바뀌었다(seed 파싱 {expected}건) — 기대값 재확인 필요"

    h = _headers(client)
    resp = _upload(client, h, XLSX.read_bytes(), name=XLSX.name)
    assert resp.status_code == 200, resp.text
    assert resp.json()["summary"]["valid_rows"] == expected
    assert resp.json()["summary"]["errors"] == []
